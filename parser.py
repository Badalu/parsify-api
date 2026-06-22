"""
parser.py — Parsify.in Hybrid PDF Parser
=========================================
Production-ready bank statement parser for Indian Chartered Accountants.

Parsing flow (cost-optimized):
  1. check_pdf_basic()       — read first 2 pages, detect text vs scanned
  2. SCANNED → Gemini File API (1 API call, handles OCR)
  3. TEXT-BASED → native pdfplumber table extraction (0 API calls)
  4. Native fails → line-by-line regex parser (0 API calls)
  5. Native < 5 txns → is_text_quality_sufficient()
  6. Quality OK → Gemini text chunks (1-2 API calls)
  7. Quality poor → Gemini File API (1 API call)

SDK: Prefers google.genai (new SDK), falls back to google.generativeai (legacy).
"""

import os
import re
import json
import time
import hashlib
import pdfplumber
from pypdf import PdfReader
import pandas as pd
from typing import List, Dict, Any, Tuple, Optional
from pydantic import BaseModel, Field

import importlib.util as _ilu

# ── SDK Detection ─────────────────────────────────────────────────────────────
# Prefer the new google-genai SDK; fall back to the old google-generativeai
if _ilu.find_spec("google.genai") is not None:
    from google import genai
    from google.genai import types as genai_types
    GENAI_NEW_SDK = True
else:
    import google.generativeai as genai  # type: ignore[no-redef]
    genai_types = None
    GENAI_NEW_SDK = False

# ── Configuration ─────────────────────────────────────────────────────────────
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

if GEMINI_API_KEY:
    if GEMINI_API_KEY.startswith("AIza"):
        print(f"[Config] Gemini API Key loaded (valid format, ends with ...{GEMINI_API_KEY[-4:]})")
    else:
        print(f"[Config] WARNING: Gemini API Key may be INVALID — expected to start with 'AIza', got '{GEMINI_API_KEY[:4]}...'")
else:
    print("[Config] WARNING: GEMINI_API_KEY not set — AI parsing will be unavailable")

# Larger chunks = fewer API calls = lower cost
GEMINI_CHUNK_SIZE = 40_000

# ── Pydantic Schemas for Structured Gemini Output ─────────────────────────────

class TransactionItem(BaseModel):
    date: str = Field(description="Transaction date exactly as shown in the statement")
    value_date: str = Field(description="Value/clearing date if listed separately, else same as date")
    description: str = Field(description="Full transaction narration/particulars")
    debit: str = Field(description="Withdrawal amount, digits only, empty if none")
    credit: str = Field(description="Deposit amount, digits only, empty if none")
    balance: str = Field(description="Running balance after transaction, digits only, empty if none")
    category: str = Field(description="Category: Salary|Food|Fuel|Rent|Utilities|Shopping|Groceries|Transfer|Subscription|GST|Tax|Investment|EMI|Refund|Other")
    gst: str = Field(description="CGST+SGST | IGST | GST | empty string")


class TransactionsList(BaseModel):
    transactions: List[TransactionItem]


class CategorizationItem(BaseModel):
    index: int = Field(description="Index of the transaction in the input list")
    category: str = Field(description="Category: Salary|Food|Fuel|Rent|Utilities|Shopping|Groceries|Transfer|Subscription|GST|Tax|Investment|EMI|Refund|Other")
    gst: str = Field(description="CGST+SGST | IGST | GST | empty string")


class CategorizationList(BaseModel):
    items: List[CategorizationItem]


# ══════════════════════════════════════════════════════════════════════════════
# SYSTEM PROMPTS — Minimal tokens (~350 tokens max), cost-optimized
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT_PARSE_ONLY = (
    "Extract every transaction from the bank statement. Never skip, never hallucinate.\n"
    "Fields: date (exactly as shown), value_date (if separate column, else copy date), "
    "description (full narration including UPI IDs/refs), "
    "debit (digits only, empty if none), credit (digits only, empty if none), "
    "balance (digits only, empty if none), category=Other, gst=empty string.\n"
    "Skip: headers, footers, opening/closing balance summary rows, blank rows.\n"
    "Bank columns: SBI=Txn Date|Value Date|Description|Debit|Credit|Balance, "
    "HDFC=Date|Narration|Value Dt|Withdrawal Amt|Deposit Amt|Closing Balance, "
    "ICICI=Transaction Date|Value Date|Particulars|Deposits|Withdrawals|Balance, "
    "Axis=Tran Date|Particulars|Debit|Credit|Balance, "
    "Kotak=Date|Description|Debit|Credit|Balance, "
    "IDFC/IndusInd/Yes Bank=compact Dr/Cr style.\n"
    "Amounts: strip ₹ Rs. INR $ symbols, strip comma separators, strip Dr/Cr suffix. "
    "Multi-line descriptions: concatenate with space."
)

SYSTEM_PROMPT_WITH_CATEGORIES = (
    "Extract every transaction from the bank statement. Never skip, never hallucinate.\n"
    "Fields: date (exactly as shown), value_date (if separate column, else copy date), "
    "description (full narration including UPI IDs/refs), "
    "debit (digits only, empty if none), credit (digits only, empty if none), "
    "balance (digits only, empty if none), category (from list below), gst (see below).\n"
    "Skip: headers, footers, opening/closing balance summary rows, blank rows.\n"
    "Bank columns: SBI=Txn Date|Value Date|Description|Debit|Credit|Balance, "
    "HDFC=Date|Narration|Value Dt|Withdrawal Amt|Deposit Amt|Closing Balance, "
    "ICICI=Transaction Date|Value Date|Particulars|Deposits|Withdrawals|Balance, "
    "Axis=Tran Date|Particulars|Debit|Credit|Balance, "
    "Kotak=Date|Description|Debit|Credit|Balance, "
    "IDFC/IndusInd/Yes Bank=compact Dr/Cr style.\n"
    "Amounts: strip ₹ Rs. INR $ symbols, strip comma separators, strip Dr/Cr suffix.\n"
    "Categories (pick exactly one): "
    "Salary=payroll/NEFT salary, Food=Zomato/Swiggy/restaurant, "
    "Fuel=petrol/diesel/HPCL/BPCL, Rent=house rent/PG, "
    "Utilities=electricity/Jio/Airtel, Shopping=Amazon/Flipkart/Myntra, "
    "Groceries=Blinkit/Zepto/BigBasket, "
    "Transfer=UPI/NEFT/RTGS/IMPS/GPay/PhonePe, "
    "Subscription=Netflix/Spotify/Prime, GST=CGST/SGST/IGST payments, "
    "Tax=income tax/TDS, Investment=Zerodha/Groww/SIP/MF, "
    "EMI=loan EMI/home loan/car loan, Refund=refund/cashback/reversal, "
    "Other=unclassified.\n"
    "GST field: CGST+SGST if both mentioned | IGST if IGST mentioned | GST if generic | empty string otherwise."
)

SYSTEM_PROMPT_CATEGORIZE = (
    "Input: list of {index, description}. Output: list of {index, category, gst}.\n"
    "Categories (pick exactly one): "
    "Salary=payroll/NEFT salary, Food=Zomato/Swiggy/restaurant, "
    "Fuel=petrol/diesel/HPCL/BPCL, Rent=house rent/PG, "
    "Utilities=electricity/Jio/Airtel, Shopping=Amazon/Flipkart/Myntra, "
    "Groceries=Blinkit/Zepto/BigBasket, "
    "Transfer=UPI/NEFT/RTGS/IMPS/GPay/PhonePe, "
    "Subscription=Netflix/Spotify/Prime, GST=CGST/SGST/IGST payments, "
    "Tax=income tax/TDS, Investment=Zerodha/Groww/SIP/MF, "
    "EMI=loan EMI/home loan/car loan, Refund=refund/cashback/reversal, "
    "Other=unclassified.\n"
    "GST field: CGST+SGST | IGST | GST | empty string."
)


# ══════════════════════════════════════════════════════════════════════════════
# PDF BASIC CHECK
# ══════════════════════════════════════════════════════════════════════════════

def check_pdf_basic(pdf_path: str, password: str = None) -> Tuple[int, bool]:
    """
    Read first 2 pages only, detect if text-based or scanned.
    Returns (total_pages, is_text_based).
    """
    try:
        reader = PdfReader(pdf_path, password=password)
        total_pages = len(reader.pages)
        if total_pages == 0:
            return 0, False

        # Check up to 2 pages for extractable text
        for i in range(min(2, total_pages)):
            text = reader.pages[i].extract_text()
            if text and len(text.strip()) > 20:
                return total_pages, True
        return total_pages, False
    except Exception as e:
        raise e


def is_text_quality_sufficient(text: str) -> bool:
    """
    Check if extracted text has enough quality for Gemini text-chunk parsing.
    Returns False if text is garbled, mostly whitespace, or encoding garbage.
    """
    if not text or len(text.strip()) < 100:
        return False

    # Check ratio of printable ASCII + common Unicode to total chars
    printable_count = 0
    total = len(text)
    for ch in text:
        if ch.isprintable() or ch in ('\n', '\t', '\r'):
            printable_count += 1

    printable_ratio = printable_count / total if total > 0 else 0
    if printable_ratio < 0.85:
        print(f"[TextQuality] Printable ratio too low: {printable_ratio:.2f}")
        return False

    # Check for presence of date-like patterns (bank statements always have dates)
    date_matches = re.findall(
        r'\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4}|'
        r'\d{1,2}\s+[A-Za-z]{3,4}\s+\d{2,4}|'
        r'\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}',
        text[:5000]
    )
    if len(date_matches) < 2:
        print(f"[TextQuality] Too few date patterns found: {len(date_matches)}")
        return False

    # Check for amount-like patterns (digits with optional comma/decimal)
    amount_matches = re.findall(r'\d{1,3}(?:,\d{2,3})*(?:\.\d{1,2})?', text[:5000])
    if len(amount_matches) < 3:
        print(f"[TextQuality] Too few amount patterns found: {len(amount_matches)}")
        return False

    # Check for garbage sequences (consecutive non-ASCII or control chars)
    garbage_runs = re.findall(r'[^\x20-\x7E\n\t\r₹]{10,}', text[:5000])
    if len(garbage_runs) > 5:
        print(f"[TextQuality] Too many garbage runs: {len(garbage_runs)}")
        return False

    return True


def extract_full_text(pdf_path: str, password: str = None) -> str:
    """
    Extract full text from all pages using pypdf (fast),
    with fallback to pdfplumber layout-aware extraction.
    """
    try:
        reader = PdfReader(pdf_path, password=password)
        extracted_pages = []
        for page in reader.pages:
            extracted_pages.append(page.extract_text() or "")
        return "\n\n--- Page Break ---\n\n".join(extracted_pages)
    except Exception as e:
        print(f"Error during pypdf text extraction: {e} — falling back to pdfplumber")
        extracted_pages = []
        try:
            with pdfplumber.open(pdf_path, password=password) as pdf:
                for page in pdf.pages:
                    try:
                        text = page.extract_text(layout=True)
                        if text:
                            extracted_pages.append(text)
                    except Exception:
                        pass
                    finally:
                        page.flush_cache()
            return "\n\n--- Page Break ---\n\n".join(extracted_pages)
        except Exception as e2:
            print(f"Fallback pdfplumber text extraction failed: {e2}")
            return ""


# ══════════════════════════════════════════════════════════════════════════════
# NATIVE RULE-BASED PARSER
# ══════════════════════════════════════════════════════════════════════════════

DATE_PATTERNS = [
    re.compile(r'^\s*\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4}\s*$'),          # 12/05/2023, 12-05-23
    re.compile(r'^\s*\d{1,2}[-/\.][A-Za-z]{3,9}[-/\.]\d{2,4}\s*$'),   # 12-May-2023, 12-September-23
    re.compile(r'^\s*\d{1,2}\s*[A-Za-z]{3,9}\s*\d{2,4}\s*$'),          # 12May2023, 12 May 23, 12May23
    re.compile(r'^\s*\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4}\s*$'),          # 12 May 2023, 12 May 23
    re.compile(r'^\s*[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{2,4}\s*$'),       # May 12, 2023, May 12 23
    re.compile(r'^\s*\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}\s*$'),            # 2023-05-12
    re.compile(r'^\s*\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\s*$'),              # 12/05/23 (no dot)
    re.compile(r'^\s*\d{1,2}\s*[-/\.]\s*\d{1,2}\s*[-/\.]\s*\d{2,4}\s*$'),  # 12 / 05 / 2023 (spaced)
    re.compile(r'^\s*\d{2}\d{2}\d{4}\s*$'),                             # 12052023 (compact, 8 digits)
    re.compile(r'^\s*\d{1,2}[-/\.\s][A-Za-z]{3,9}[-/\.\s]?\d{0,4}\s*$'),  # 12-May or 12 May (no year)
]

# Pre-compiled cleanup for date values
_DATE_CLEANUP_RE = re.compile(r'[\n\r\t]+')


def is_valid_date(val: str) -> bool:
    """Regex check for DD/MM/YYYY, DD-MM-YY, DD Mon YYYY, YYYY-MM-DD etc."""
    if not val:
        return False
    val_str = str(val).strip()
    # Clean up newlines/tabs that pdfplumber sometimes leaves in cell values
    val_str = _DATE_CLEANUP_RE.sub(' ', val_str).strip()
    if not val_str or len(val_str) < 4 or len(val_str) > 25:
        return False
    for pattern in DATE_PATTERNS:
        if pattern.match(val_str):
            return True
    return False


# ── Local Categorization (FREE — no API calls) ───────────────────────────────

def categorize_locally(description: str) -> str:
    """Keyword matching for transaction categorization. Returns one category."""
    desc_lower = str(description).lower()

    categories = {
        "Salary": ["salary", "payroll", "wage", "direct dep", "neft salary"],
        "Food": ["zomato", "swiggy", "restaurant", "cafe", "food", "dining", "domino", "starbuck"],
        "Fuel": ["petrol", "diesel", "hpcl", "bpcl", "iocl", "fuel", "shell"],
        "Rent": ["rent", "landlord", "brokerage", "house rent", "pg rent"],
        "Utilities": [
            "electricity", "bescom", "tneb", "uppcl", "water", "broadband",
            "jio", "airtel", "vi ", "recharge", "gas", "indane", "act fibernet"
        ],
        "Shopping": ["amazon", "flipkart", "myntra", "meesho", "clothing", "appliances", "retail"],
        "Groceries": [
            "blinkit", "zepto", "instamart", "bigbasket", "supermarket",
            "grocery", "groceries", "milk", "dairy"
        ],
        "Transfer": [
            "upi", "transfer", "neft", "rtgs", "imps", "ft",
            "sent to", "received from", "gpay", "phonepe"
        ],
        "Subscription": [
            "netflix", "prime video", "spotify", "youtube premium",
            "disney", "hotstar", "microsoft", "google storage", "cloud"
        ],
        "GST": ["gst", "cgst", "sgst", "igst", "tax split"],
        "Tax": ["income tax", "tds", "itr", "advance tax", "professional tax"],
        "Investment": ["zerodha", "groww", "mutual fund", "sip", "demat", "shares", "stocks", "etf"],
        "EMI": ["emi", "loan", "mortgage", "hdfc bank loan", "sbi loan", "car emi", "home loan"],
        "Refund": ["refund", "cashback", "returned", "reversal"],
    }

    for cat, keywords in categories.items():
        if any(kw in desc_lower for kw in keywords):
            return cat

    return "Other"


def extract_gst_locally(description: str) -> str:
    """Returns CGST+SGST / IGST / GST / empty string."""
    desc_lower = str(description).lower()
    if "cgst" in desc_lower and "sgst" in desc_lower:
        return "CGST+SGST"
    elif "igst" in desc_lower:
        return "IGST"
    elif "gst" in desc_lower:
        return "GST"
    return ""


# ── Header Detection ─────────────────────────────────────────────────────────

def find_header_mapping(row: List[str]) -> Dict[str, int]:
    """
    Detect column positions from header row keywords.
    Date/Dt/Txn Date → date col, Value Date/Val → value_date col,
    Particulars/Narration/Description/Remarks → description col,
    Debit/Withdrawal/Dr → debit col, Credit/Deposit/Cr → credit col,
    Balance/Bal → balance col.
    """
    mapping = {}
    row_lower = [str(cell).lower().strip() if cell is not None else "" for cell in row]
    # Also clean newlines within cell text (pdfplumber sometimes has these)
    row_lower = [re.sub(r'[\n\r\t]+', ' ', c).strip() for c in row_lower]

    date_kws = ["date", "dt", "txn d", "trans d", "posting d", "post d", "trn d"]
    desc_kws = [
        "particulars", "description", "narration", "remarks",
        "transaction details", "details", "narrative", "desc",
        "party name", "beneficiary", "payee"
    ]
    debit_kws = ["debit", "withdrawal", "payment", "withdraw", "withdrawal(dr)", "withdrawal amt", "dr amount"]
    credit_kws = ["credit", "deposit", "receipt", "deposit(cr)", "deposit amt", "cr amount"]
    balance_kws = ["balance", "bal", "running", "closing bal", "available"]
    amount_kws = ["amount", "amt", "transaction value", "txn amt", "txn amount", "transaction amt"]
    dr_cr_kws = ["dr/cr", "dr / cr", "d/c", "dr_cr", "cr/dr", "cr / dr", "dr./cr."]

    for idx, cell in enumerate(row_lower):
        if not cell:
            continue

        # Date columns
        if any(kw in cell for kw in date_kws):
            if "value" in cell or "val" in cell or "clearing" in cell:
                mapping["value_date"] = idx
            else:
                if "date" not in mapping:  # take the first date column found
                    mapping["date"] = idx

        # Description columns
        if any(kw in cell for kw in desc_kws):
            if "description" not in mapping:  # take the first description column
                mapping["description"] = idx

        # Debit columns — but NOT if the cell is just "dr" or "cr" alone (that's dr_cr column)
        if any(kw in cell for kw in debit_kws):
            if cell not in ("dr", "cr", "d", "c", "dr/cr", "cr/dr", "type", "d/c"):
                mapping["debit"] = idx

        # Credit columns
        if any(kw in cell for kw in credit_kws):
            if cell not in ("dr", "cr", "d", "c", "dr/cr", "cr/dr", "type", "d/c"):
                mapping["credit"] = idx

        # Single Amount column
        if any(kw in cell for kw in amount_kws):
            if not any(skip in cell for skip in ["debit", "credit", "payment", "deposit", "withdraw", "dr", "cr"]):
                mapping["amount"] = idx

        # Dr/Cr type column
        if any(kw in cell for kw in dr_cr_kws) or cell in ["dr", "cr", "d/c", "type"]:
            mapping["dr_cr"] = idx

        # Balance columns
        if any(kw in cell for kw in balance_kws):
            mapping["balance"] = idx

    # Fallback if date is not found but value_date is
    if "date" not in mapping and "value_date" in mapping:
        mapping["date"] = mapping["value_date"]

    # Accept if we have at least date + description OR date + amount
    if "date" in mapping and ("description" in mapping or "amount" in mapping):
        return mapping
    return {}


# ── Line-by-Line Regex Parser ────────────────────────────────────────────────

def _is_amount_str(s: str) -> bool:
    """Check if a string looks like a monetary amount (handles Indian format)."""
    if not s:
        return False
    s_clean = s.strip()
    # Remove currency symbols
    s_clean = re.sub(r'(?i)(rs\.?|inr|usd|₹|\$|£|€)', '', s_clean).strip()
    # Handle parenthetical negatives: (500.00) → 500.00
    if s_clean.startswith('(') and s_clean.endswith(')'):
        s_clean = s_clean[1:-1].strip()
    # Remove Dr/Cr suffixes
    s_clean = re.sub(r'(?i)\s*(dr|cr)\s*$', '', s_clean).strip()
    # Remove commas, plus, minus, spaces
    s_clean = s_clean.replace(",", "").replace("+", "").replace(" ", "").strip()
    # Remove leading minus
    if s_clean.startswith("-"):
        s_clean = s_clean[1:]
    if not s_clean:
        return False
    try:
        float(s_clean)
        return True
    except ValueError:
        return False


def _has_dr_suffix(s: str) -> bool:
    """Check if amount has Dr suffix (debit indicator)."""
    return bool(re.search(r'(?i)\s*dr\s*$', s.strip()))


def _has_cr_suffix(s: str) -> bool:
    """Check if amount has Cr suffix (credit indicator)."""
    return bool(re.search(r'(?i)\s*cr\s*$', s.strip()))


def parse_line_by_line(text: str) -> List[Dict[str, Any]]:
    """
    Fallback parser that extracts transactions from raw page text line-by-line.
    Handles borderless table layouts common in Indian bank statements.
    """
    transactions = []
    lines = text.split("\n")

    header_bounds = {}

    date_kws = ["date", "dt", "txn d", "trans d", "posting d", "trn d"]
    desc_kws = [
        "particulars", "description", "narration", "remarks",
        "transaction details", "details", "narrative", "desc"
    ]
    debit_kws = ["debit", "withdrawal", "payment", "dr", "withdraw", "withdrawal(dr)", "withdrawal amt"]
    credit_kws = ["credit", "deposit", "receipt", "cr", "deposit(cr)", "deposit amt"]
    balance_kws = ["balance", "bal", "running", "closing"]

    # Skip words for non-transaction lines
    skip_words = [
        "balance brought", "carried forward", "balance carried", "brought forward",
        "total", "page ", "statement", "generated on", "opening balance",
        "closing balance", "printed on", "branch", "customer", "account no",
        "ifsc", "micr", "address", "email", "mobile", "phone"
    ]

    # First pass: find the header line and compute column boundaries
    for line in lines:
        line_lower = line.lower()
        if any(kw in line_lower for kw in date_kws) and any(kw in line_lower for kw in desc_kws):
            bounds = {}

            # Find date columns — be careful about value_date vs date
            for kw in date_kws:
                idx = line_lower.find(kw)
                if idx != -1:
                    surrounding = line_lower[max(0, idx - 12):idx + 25]
                    if any(v in surrounding for v in ["value", "val ", "clearing", "post"]):
                        bounds["value_date"] = idx
                    elif "date" not in bounds:
                        bounds["date"] = idx

            if "date" not in bounds:
                for kw in date_kws:
                    idx = line_lower.find(kw)
                    if idx != -1:
                        bounds["date"] = idx
                        break

            for kw in desc_kws:
                idx = line_lower.find(kw)
                if idx != -1:
                    bounds["description"] = idx
                    break

            for kw in debit_kws:
                idx = line_lower.find(kw)
                if idx != -1:
                    bounds["debit"] = idx
                    break

            for kw in credit_kws:
                idx = line_lower.find(kw)
                if idx != -1:
                    bounds["credit"] = idx
                    break

            for kw in balance_kws:
                idx = line_lower.find(kw)
                if idx != -1:
                    # Make sure it's not "closing balance" used as a summary
                    if "closing" in line_lower[max(0, idx - 10):idx + 2]:
                        # Only use if there are other column headers too
                        if len(bounds) >= 2:
                            bounds["balance"] = idx
                    else:
                        bounds["balance"] = idx
                    break

            if "date" in bounds and "description" in bounds:
                sorted_cols = sorted(bounds.items(), key=lambda x: x[1])
                header_bounds = {}
                for i in range(len(sorted_cols)):
                    col_name, start = sorted_cols[i]
                    # Use the midpoint between this column's start and the previous column's start
                    if i > 0:
                        prev_name, prev_start = sorted_cols[i - 1]
                        col_start = start - 2  # start 2 chars before the keyword
                    else:
                        col_start = 0

                    if i < len(sorted_cols) - 1:
                        next_start = sorted_cols[i + 1][1]
                        col_end = next_start - 1
                    else:
                        col_end = max(500, len(line) + 50)  # extend for last column

                    header_bounds[col_name] = (max(0, col_start), col_end)
                break

    # Second pass: extract transactions using column boundaries
    for line in lines:
        if not line.strip():
            continue

        line_lower = line.lower().strip()

        # Skip header lines
        if any(kw in line_lower for kw in date_kws) and any(kw in line_lower for kw in desc_kws):
            continue

        # Skip non-transaction footer/summary lines
        if any(sw in line_lower for sw in skip_words):
            continue

        if header_bounds:
            date_b = header_bounds.get("date")
            desc_b = header_bounds.get("description")
            debit_b = header_bounds.get("debit")
            credit_b = header_bounds.get("credit")
            balance_b = header_bounds.get("balance")
            val_date_b = header_bounds.get("value_date")

            def get_slice(b):
                if not b:
                    return ""
                s, e = b
                return line[s:min(e, len(line))].strip() if s < len(line) else ""

            date_val = get_slice(date_b)
            desc_val = get_slice(desc_b)

            if is_valid_date(date_val):
                debit_val = get_slice(debit_b)
                credit_val = get_slice(credit_b)
                balance_val = get_slice(balance_b)
                val_date_val = get_slice(val_date_b)

                if not desc_val:
                    continue

                transactions.append({
                    "date": date_val,
                    "value_date": val_date_val if val_date_val else date_val,
                    "description": desc_val,
                    "debit": debit_val,
                    "credit": credit_val,
                    "balance": balance_val,
                    "category": categorize_locally(desc_val),
                    "gst": extract_gst_locally(desc_val),
                })
                continue

            elif desc_val and transactions:
                if not any(sw in desc_val.lower() for sw in skip_words):
                    transactions[-1]["description"] += " " + desc_val

                    debit_val = get_slice(debit_b)
                    credit_val = get_slice(credit_b)
                    balance_val = get_slice(balance_b)

                    if debit_val and not transactions[-1]["debit"]:
                        transactions[-1]["debit"] = debit_val
                    if credit_val and not transactions[-1]["credit"]:
                        transactions[-1]["credit"] = credit_val
                    if balance_val and not transactions[-1]["balance"]:
                        transactions[-1]["balance"] = balance_val
                continue

        # Fallback: split by multiple spaces when no header bounds
        parts = re.split(r'\s{2,}', line.strip())
        if not parts:
            continue

        date_val = ""
        date_part_idx = -1
        for idx, p in enumerate(parts):
            if is_valid_date(p):
                date_val = p
                date_part_idx = idx
                break

        if date_val and date_part_idx != -1:
            remaining = parts[date_part_idx + 1:]
            if not remaining:
                continue

            value_date_val = date_val
            if remaining and is_valid_date(remaining[0]):
                value_date_val = remaining.pop(0)

            if not remaining:
                continue

            amounts = []
            desc_parts = []

            for r in remaining:
                if _is_amount_str(r):
                    amounts.append(r)
                else:
                    desc_parts.append(r)

            desc_val = " ".join(desc_parts).strip()
            if not desc_val:
                # If no desc but amounts exist, skip (probably a summary line)
                continue

            debit_val = ""
            credit_val = ""
            balance_val = ""

            if len(amounts) == 1:
                amt = amounts[0]
                # Check for inline Dr/Cr suffix
                if _has_dr_suffix(amt) or amt.startswith("-") or amt.startswith("("):
                    debit_val = amt
                elif _has_cr_suffix(amt):
                    credit_val = amt
                else:
                    # Guess from description
                    desc_lower = desc_val.lower()
                    is_deb = any(kw in desc_lower for kw in [
                        "payment", "upi to", "transfer to", "debit", "withdrawal",
                        "paid to", "neft to", "rtgs to", "imps to", "emi", "loan"
                    ])
                    if is_deb:
                        debit_val = amt
                    else:
                        credit_val = amt
            elif len(amounts) == 2:
                amt = amounts[0]
                bal = amounts[1]
                # Check inline Dr/Cr
                if _has_dr_suffix(amt) or amt.startswith("-") or amt.startswith("("):
                    debit_val = amt
                elif _has_cr_suffix(amt):
                    credit_val = amt
                else:
                    desc_lower = desc_val.lower()
                    is_deb = any(kw in desc_lower for kw in [
                        "payment", "upi to", "transfer to", "debit", "withdrawal",
                        "paid to", "neft to", "rtgs to", "imps to", "emi", "loan"
                    ]) or amt.startswith("-")
                    if is_deb:
                        debit_val = amt
                    else:
                        credit_val = amt
                balance_val = bal
            elif len(amounts) >= 3:
                debit_val = amounts[0]
                credit_val = amounts[1]
                balance_val = amounts[2]

            transactions.append({
                "date": date_val,
                "value_date": value_date_val,
                "description": desc_val,
                "debit": debit_val,
                "credit": credit_val,
                "balance": balance_val,
                "category": categorize_locally(desc_val),
                "gst": extract_gst_locally(desc_val),
            })

        elif transactions and len(parts) <= 2:
            # Continuation line (multi-line description or trailing amounts)
            val = " ".join(parts).strip()
            if not any(sw in val.lower() for sw in skip_words) and len(val) > 3:
                # Check if this is a trailing amount for the last txn
                if len(parts) == 1 and _is_amount_str(val):
                    if not transactions[-1]["balance"]:
                        transactions[-1]["balance"] = val
                    elif not transactions[-1]["debit"] and not transactions[-1]["credit"]:
                        transactions[-1]["credit"] = val  # default to credit
                else:
                    transactions[-1]["description"] += " " + val

    return transactions


# ── Native pdfplumber Table Parser ────────────────────────────────────────────

def parse_pdf_natively(pdf_path: str, password: str = None) -> List[Dict[str, Any]]:
    """
    Extract transactions from structured PDF statements natively using pdfplumber.
    30s timeout, 150 page limit. Includes:
    - Standard table extraction + borderless text-strategy fallback
    - Column-shift repair for borderless tables
    - Repeated header detection (skip headers on every page)
    - Cell value cleanup (newlines, whitespace)
    """
    start_time = time.time()
    max_pages = 150

    try:
        reader = PdfReader(pdf_path, password=password)
        total_pages = len(reader.pages)
        if total_pages > max_pages:
            print(f"[Native] PDF has {total_pages} pages (limit: {max_pages}). Bypassing native parsing.")
            return []
    except Exception as e:
        print(f"[Native] Failed to check page count via pypdf: {e}")

    transactions: List[Dict[str, Any]] = []
    header_mapping: Dict[str, int] = {}

    # Table settings to try: default first, then text-based for borderless tables
    table_settings_list = [
        {},  # default pdfplumber settings
        {"vertical_strategy": "text", "horizontal_strategy": "text"},  # borderless
    ]

    try:
        with pdfplumber.open(pdf_path, password=password) as pdf:
            table_found = False
            for idx, page in enumerate(pdf.pages):
                # 30s timeout guard
                if time.time() - start_time > 30.0:
                    print(f"[Native] Timeout of 30s exceeded at page {idx + 1}/{len(pdf.pages)}. Returning what we have.")
                    break

                if idx >= 3 and not table_found:
                    print("[Native] No tables found in first 3 pages. Aborting table extraction.")
                    break

                # Try multiple table extraction strategies
                tables = []
                for settings in table_settings_list:
                    try:
                        if settings:
                            extracted = page.extract_tables(table_settings=settings)
                        else:
                            extracted = page.extract_tables()
                        if extracted:
                            tables = extracted
                            break
                    except Exception:
                        continue

                # Fallback to find_tables
                if not tables:
                    try:
                        found = page.find_tables()
                        if found:
                            tables = [t.extract() for t in found]
                    except Exception:
                        pass

                if tables:
                    table_found = True

                for table in tables:
                    if not table:
                        continue

                    for row in table:
                        if row is None:
                            continue
                        # Clean cells: handle None, strip whitespace, clean newlines
                        clean_row = []
                        for cell in row:
                            if cell is None:
                                clean_row.append("")
                            else:
                                cleaned = str(cell).strip()
                                cleaned = re.sub(r'[\n\r\t]+', ' ', cleaned).strip()
                                clean_row.append(cleaned)

                        if not any(clean_row):
                            continue

                        # Check if this row is a header (could be repeated on each page)
                        mapping = find_header_mapping(clean_row)
                        if mapping:
                            header_mapping = mapping
                            continue

                        if not header_mapping:
                            continue

                        date_idx = header_mapping.get("date")
                        desc_idx = header_mapping.get("description")
                        debit_idx = header_mapping.get("debit")
                        credit_idx = header_mapping.get("credit")
                        amount_idx = header_mapping.get("amount")
                        dr_cr_idx = header_mapping.get("dr_cr")
                        balance_idx = header_mapping.get("balance")
                        val_date_idx = header_mapping.get("value_date")

                        def _safe_get(idx_val, row_data):
                            if idx_val is not None and idx_val < len(row_data):
                                return row_data[idx_val]
                            return ""

                        date_val = _safe_get(date_idx, clean_row)
                        desc_val = _safe_get(desc_idx, clean_row)

                        if not desc_val and not date_val:
                            continue

                        # Skip summary/footer rows
                        combined_lower = (date_val + " " + desc_val).lower()
                        if any(sw in combined_lower for sw in [
                            "opening balance", "closing balance", "balance brought",
                            "balance carried", "total", "grand total", "statement generated"
                        ]):
                            continue

                        if is_valid_date(date_val):
                            debit_val = _safe_get(debit_idx, clean_row)
                            credit_val = _safe_get(credit_idx, clean_row)
                            
                            # Handle single amount column + dr/cr column
                            if not debit_val and not credit_val and amount_idx is not None:
                                amt_val = _safe_get(amount_idx, clean_row)
                                type_val = _safe_get(dr_cr_idx, clean_row).lower().strip()
                                
                                if "dr" in type_val or type_val == "d":
                                    debit_val = amt_val
                                elif "cr" in type_val or type_val == "c":
                                    credit_val = amt_val
                                else:
                                    # Fallback: check amount sign or inline dr/cr
                                    amt_lower = amt_val.lower() if amt_val else ""
                                    if amt_val and amt_val.startswith("-"):
                                        debit_val = amt_val.lstrip("-")
                                    elif "dr" in amt_lower:
                                        debit_val = re.sub(r'(?i)\s*dr\s*$', '', amt_val)
                                    elif "cr" in amt_lower:
                                        credit_val = re.sub(r'(?i)\s*cr\s*$', '', amt_val)
                                    elif amt_val and amt_val.startswith("(") and amt_val.endswith(")"):
                                        debit_val = amt_val[1:-1]
                                    else:
                                        # Guess from description
                                        if desc_val and any(kw in desc_val.lower() for kw in [
                                            "payment", "withdrawal", "debit", "upi to", "neft to",
                                            "paid", "emi", "loan"
                                        ]):
                                            debit_val = amt_val
                                        else:
                                            credit_val = amt_val

                            balance_val = _safe_get(balance_idx, clean_row)
                            val_date_val = _safe_get(val_date_idx, clean_row)

                            # Repair column shift: when value_date col has text instead of date (borderless tables)
                            if amount_idx is None:
                                if (val_date_idx is not None and val_date_val
                                        and len(val_date_val) > 11
                                        and not is_valid_date(val_date_val)
                                        and re.search(r'[A-Za-z]{2,}', val_date_val)):
                                    balance_val = _safe_get(credit_idx, clean_row)
                                    credit_val = _safe_get(debit_idx, clean_row)
                                    debit_val = _safe_get(desc_idx, clean_row)
                                    desc_val = val_date_val
                                    val_date_val = date_val

                                # If debit is still alphabetic (not an amount), description spilled over
                                if debit_val and re.search(r'[A-Za-z]{3,}', debit_val) and "dr" not in debit_val.lower():
                                    desc_val += " " + debit_val
                                    debit_val = credit_val
                                    credit_val = balance_val
                                    balance_val = clean_row[balance_idx + 1] if balance_idx is not None and balance_idx + 1 < len(clean_row) else ""

                            txn = {
                                "date": date_val,
                                "value_date": val_date_val if val_date_val else date_val,
                                "description": desc_val,
                                "debit": debit_val,
                                "credit": credit_val,
                                "balance": balance_val,
                                "category": categorize_locally(desc_val),
                                "gst": extract_gst_locally(desc_val),
                            }
                            transactions.append(txn)

                        elif desc_val and not date_val and transactions:
                            desc_lower = desc_val.lower()
                            skip_words = ["balance", "carried", "brought", "total", "page", "statement"]
                            if not any(sw in desc_lower for sw in skip_words):
                                transactions[-1]["description"] += " " + desc_val

                                if debit_idx is not None and debit_idx < len(clean_row) and clean_row[debit_idx] and not transactions[-1]["debit"]:
                                    transactions[-1]["debit"] = clean_row[debit_idx]
                                if credit_idx is not None and credit_idx < len(clean_row) and clean_row[credit_idx] and not transactions[-1]["credit"]:
                                    transactions[-1]["credit"] = clean_row[credit_idx]
                                if balance_idx is not None and balance_idx < len(clean_row) and clean_row[balance_idx] and not transactions[-1]["balance"]:
                                    transactions[-1]["balance"] = clean_row[balance_idx]

                page.flush_cache()

        # If table extraction yielded nothing, try line-by-line regex fallback
        if not transactions:
            print("[Native] No transactions from table extraction. Trying line-by-line regex fallback...")
            full_text_list = []
            with pdfplumber.open(pdf_path, password=password) as pdf:
                for page in pdf.pages:
                    try:
                        page_text = page.extract_text(layout=True)
                        if page_text:
                            full_text_list.append(page_text)
                    except Exception:
                        pass
                    finally:
                        page.flush_cache()
            if full_text_list:
                raw_text = "\n".join(full_text_list)
                transactions = parse_line_by_line(raw_text)

    except Exception as e:
        print(f"Error during native PDF parsing: {e}")

    return transactions


# ══════════════════════════════════════════════════════════════════════════════
# GEMINI AI PARSING
# ══════════════════════════════════════════════════════════════════════════════

def _get_system_prompt(categorize: bool, gst: bool) -> str:
    """Select the appropriate minimal system prompt based on flags."""
    if categorize:
        return SYSTEM_PROMPT_WITH_CATEGORIES
    else:
        return SYSTEM_PROMPT_PARSE_ONLY


def _strip_markdown_fences(raw: str) -> str:
    """Clean up potential markdown code fence wrapping from Gemini response."""
    raw = raw.strip()
    if raw.startswith("```json"):
        raw = raw[7:]
    if raw.startswith("```"):
        raw = raw[3:]
    if raw.endswith("```"):
        raw = raw[:-3]
    return raw.strip()


def _call_gemini_chunk(
    model_or_client,
    system_prompt: str,
    text_chunk: str,
    chunk_num: int = 1,
    chunk_context: str = "",
) -> List[Dict[str, Any]]:
    """
    Call Gemini with a single chunk of text, with retry + exponential backoff.
    Max 5 retries. Rate limit (429/quota): sleep 10s then retry.
    Billing/permission errors: raise immediately, do not retry.
    """
    user_message = (
        f"Extract ALL transactions from this bank statement text. "
        f"Return every transaction row, do not skip any.\n"
        f"{chunk_context}\n"
        f"BANK STATEMENT TEXT:\n{'=' * 60}\n{text_chunk}\n{'=' * 60}"
    )

    max_retries = 3
    for attempt in range(max_retries):
        try:
            if GENAI_NEW_SDK:
                client = model_or_client
                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=user_message,
                    config=genai_types.GenerateContentConfig(
                        system_instruction=system_prompt,
                        response_mime_type="application/json",
                        response_schema=TransactionsList,
                        temperature=0,
                    ),
                )
            else:
                model = model_or_client
                response = model.generate_content(
                    contents=[{"role": "user", "parts": [user_message]}],
                    generation_config=genai.GenerationConfig(
                        response_mime_type="application/json",
                        response_schema=TransactionsList,
                        temperature=0,
                    ),
                )

            raw = _strip_markdown_fences(response.text)
            data = json.loads(raw)
            txns = data.get("transactions", [])
            print(f"  Chunk {chunk_num}: Extracted {len(txns)} transactions")
            return txns
        except Exception as e:
            err_msg = str(e).lower()

            # Billing/permission errors: raise immediately
            if any(kw in err_msg for kw in ["dunning", "billing", "permission_denied", "403"]):
                print(f"  Chunk {chunk_num} billing/permission error: {e}")
                raise

            print(f"  Chunk {chunk_num} attempt {attempt + 1} failed: {e}")

            if "429" in err_msg or "quota" in err_msg or "resource_exhausted" in err_msg or "limit" in err_msg:
                print(f"Rate limit hit on chunk {chunk_num}. Sleeping 5s before retry...")
                time.sleep(5)

            if attempt < max_retries - 1:
                backoff = min(2 ** attempt, 4)  # 1s, 2s, 4s
                time.sleep(backoff)
            else:
                raise
    return []


def _deduplicate_transactions(txns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Remove exact duplicate transaction rows using MD5 fingerprint
    of date + description + debit + credit + balance.
    """
    seen = set()
    result = []
    for txn in txns:
        key = hashlib.md5(
            f"{txn.get('date', '')}__{txn.get('description', '')}__{txn.get('debit', '')}__{txn.get('credit', '')}__{txn.get('balance', '')}".encode()
        ).hexdigest()
        if key not in seen:
            seen.add(key)
            result.append(txn)
    removed = len(txns) - len(result)
    if removed:
        print(f"  Deduplication: removed {removed} exact duplicate transactions")
    return result


def parse_with_gemini(
    text: str,
    categorize: bool = False,
    gst: bool = True,
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Send statement text to Gemini and get parsed transactions.
    Handles large statements by splitting into 200K-char chunks.
    Returns (transactions_list, gemini_calls_count).
    """
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY environment variable is not configured.")

    system_prompt = _get_system_prompt(categorize, gst)
    gemini_calls = 0

    if GENAI_NEW_SDK:
        model_or_client = genai.Client(api_key=GEMINI_API_KEY)
        print("Using new google.genai SDK")
    else:
        genai.configure(api_key=GEMINI_API_KEY)
        model_or_client = genai.GenerativeModel(
            model_name="gemini-2.5-flash",
            system_instruction=system_prompt,
        )
        print("Using legacy google.generativeai SDK")

    all_transactions = []

    if len(text) <= GEMINI_CHUNK_SIZE:
        # Single call for normal-sized statements
        all_transactions = _call_gemini_chunk(model_or_client, system_prompt, text, chunk_num=1)
        gemini_calls = 1
    else:
        # Multi-chunk: split by page breaks, then by size
        print(f"Large statement ({len(text)} chars), splitting into chunks of {GEMINI_CHUNK_SIZE}...")
        pages = text.split("\n\n--- Page Break ---\n\n")

        chunks: List[Tuple[int, str]] = []
        current_chunk = ""
        chunk_num = 1

        for page in pages:
            if len(current_chunk) + len(page) > GEMINI_CHUNK_SIZE and current_chunk:
                chunks.append((chunk_num, current_chunk))
                chunk_num += 1
                current_chunk = page
            else:
                current_chunk += ("\n\n--- Page Break ---\n\n" if current_chunk else "") + page

        if current_chunk:
            chunks.append((chunk_num, current_chunk))

        total_chunks = len(chunks)
        print(f"  Processing {total_chunks} chunks in parallel...")

        from concurrent.futures import ThreadPoolExecutor

        def process_chunk(item):
            c_num, chunk_text = item
            print(f"  Processing chunk {c_num} of {total_chunks}...")
            return _call_gemini_chunk(
                model_or_client,
                system_prompt,
                chunk_text,
                c_num,
                f"(Chunk {c_num} of {total_chunks} — process only transactions in this section)",
            )

        with ThreadPoolExecutor(max_workers=min(total_chunks, 10)) as executor:
            chunk_results = list(executor.map(process_chunk, chunks))

        for txns in chunk_results:
            all_transactions.extend(txns)
            gemini_calls += 1

    # Normalize output
    result = []
    for txn in all_transactions:
        if isinstance(txn, dict):
            result.append(txn)
        else:
            result.append(txn.model_dump() if hasattr(txn, "model_dump") else dict(txn))

    # Deduplicate cross-chunk overlaps
    result = _deduplicate_transactions(result)

    print(f"Gemini total extracted: {len(result)} transactions ({gemini_calls} API calls)")
    return result, gemini_calls


# ── Gemini File API (for scanned PDFs and poor-quality text) ──────────────────

def parse_file_directly_with_gemini(
    file_path: str,
    mime_type: str,
    categorize: bool = False,
    gst: bool = True,
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Uploads a file (scanned PDF or image) to Gemini File API,
    parses with Gemini 2.5 Flash, returns (transactions, gemini_calls).
    Cleans up uploaded file in finally block always.
    """
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY environment variable is not configured.")

    system_prompt = _get_system_prompt(categorize, gst)
    user_message = (
        "Extract ALL transactions from this bank statement document/image. "
        "Return every transaction row, do not skip any."
    )

    max_retries = 2
    for attempt in range(max_retries):
        uploaded_file = None
        try:
            if GENAI_NEW_SDK:
                client = genai.Client(api_key=GEMINI_API_KEY)
                print(f"[Gemini File API] Uploading: {file_path} (type: {mime_type})")
                uploaded_file = client.files.upload(file=file_path)

                # Wait for processing (PDFs need server-side processing)
                if mime_type == "application/pdf":
                    state_str = str(uploaded_file.state).upper()
                    wait_time = 0
                    while "PROCESSING" in state_str and wait_time < 20:
                        print("Waiting for PDF processing in Files API...")
                        time.sleep(2)
                        wait_time += 2
                        uploaded_file = client.files.get(name=uploaded_file.name)
                        state_str = str(uploaded_file.state).upper()

                    if "ACTIVE" not in state_str:
                        raise Exception(f"File state is {state_str}, cannot process.")

                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=[uploaded_file, user_message],
                    config=genai_types.GenerateContentConfig(
                        system_instruction=system_prompt,
                        response_mime_type="application/json",
                        response_schema=TransactionsList,
                        temperature=0,
                    ),
                )
                raw = response.text
            else:
                genai.configure(api_key=GEMINI_API_KEY)
                model = genai.GenerativeModel(
                    model_name="gemini-2.5-flash",
                    system_instruction=system_prompt,
                )
                print(f"[Gemini File API] Uploading (legacy): {file_path} (type: {mime_type})")
                uploaded_file = genai.upload_file(file_path, mime_type=mime_type)

                if mime_type == "application/pdf":
                    state_str = str(uploaded_file.state).upper()
                    wait_time = 0
                    while "PROCESSING" in state_str and wait_time < 20:
                        print("Waiting for PDF processing in legacy Files API...")
                        time.sleep(2)
                        wait_time += 2
                        uploaded_file = genai.get_file(uploaded_file.name)
                        state_str = str(uploaded_file.state).upper()

                    if "ACTIVE" not in state_str:
                        raise Exception(f"File state is {state_str}, cannot process.")

                response = model.generate_content(
                    contents=[uploaded_file, user_message],
                    generation_config=genai.GenerationConfig(
                        response_mime_type="application/json",
                        response_schema=TransactionsList,
                        temperature=0,
                    ),
                )
                raw = response.text

            raw = _strip_markdown_fences(raw)
            data = json.loads(raw)
            txns = data.get("transactions", [])

            # Normalize
            result = []
            for txn in txns:
                if isinstance(txn, dict):
                    result.append(txn)
                else:
                    result.append(txn.model_dump() if hasattr(txn, "model_dump") else dict(txn))

            print(f"[Gemini File API] Extracted {len(result)} transactions")
            return result, 1

        except Exception as e:
            err_msg = str(e).lower()
            # Billing/permission errors: raise immediately, do not retry
            if any(kw in err_msg for kw in ["dunning", "billing", "permission_denied", "403"]):
                raise

            print(f"[Gemini File API] Attempt {attempt + 1} failed: {e}")
            if "429" in err_msg or "quota" in err_msg or "resource_exhausted" in err_msg:
                time.sleep(5)
            if attempt < max_retries - 1:
                backoff = min(2 ** attempt, 4)
                time.sleep(backoff)
            else:
                raise
        finally:
            # Always cleanup uploaded file
            if uploaded_file:
                try:
                    if GENAI_NEW_SDK:
                        cleanup_client = genai.Client(api_key=GEMINI_API_KEY)
                        cleanup_client.files.delete(name=uploaded_file.name)
                    else:
                        uploaded_file.delete()
                    print("[Gemini File API] Cleaned up uploaded file")
                except Exception as ex:
                    print(f"[Gemini File API] Cleanup failed: {ex}")

    return [], 0


# ── Gemini Categorization (post-parse enrichment) ────────────────────────────

def categorize_and_tag_with_gemini(
    transactions: List[Dict[str, Any]],
    categorize: bool = True,
    gst: bool = True,
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Categorize transactions and tag GST using Gemini 2.5 Flash in batches.
    Falls back to local rules if API call fails.
    Returns (transactions, gemini_calls).
    """
    if not transactions:
        return [], 0
    if not GEMINI_API_KEY:
        print("[Gemini] API Key not set, using local rule-based categories.")
        return transactions, 0

    system_prompt = SYSTEM_PROMPT_CATEGORIZE

    # Prepare input: only index + description to minimize token usage
    input_items = [{"index": idx, "description": txn.get("description", "")} for idx, txn in enumerate(transactions)]

    # Split into chunks of 150 descriptions
    chunk_size = 150
    chunks = [input_items[i:i + chunk_size] for i in range(0, len(input_items), chunk_size)]

    if GENAI_NEW_SDK:
        client = genai.Client(api_key=GEMINI_API_KEY)
    else:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel(model_name="gemini-2.5-flash")

    categorized_items = {}
    gemini_calls = 0

    for c_idx, chunk in enumerate(chunks):
        user_message = f"Categorize these transactions:\n{json.dumps(chunk)}"
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if GENAI_NEW_SDK:
                    response = client.models.generate_content(
                        model="gemini-2.5-flash",
                        contents=user_message,
                        config=genai_types.GenerateContentConfig(
                            system_instruction=system_prompt,
                            response_mime_type="application/json",
                            response_schema=CategorizationList,
                            temperature=0,
                        ),
                    )
                    raw = response.text
                else:
                    response = model.generate_content(
                        contents=[{"role": "user", "parts": [user_message]}],
                        generation_config=genai.GenerationConfig(
                            response_mime_type="application/json",
                            response_schema=CategorizationList,
                            temperature=0,
                        ),
                    )
                    raw = response.text

                raw = _strip_markdown_fences(raw)
                data = json.loads(raw)
                for item in data.get("items", []):
                    idx = item.get("index")
                    if idx is not None:
                        categorized_items[int(idx)] = {
                            "category": item.get("category", "Other") if categorize else "Other",
                            "gst": item.get("gst", "") if gst else "",
                        }
                gemini_calls += 1
                break  # Success
            except Exception as e:
                err_msg = str(e).lower()
                if any(kw in err_msg for kw in ["dunning", "billing", "permission_denied", "403"]):
                    raise

                print(f"[Gemini Categorize] Chunk {c_idx + 1} attempt {attempt + 1} failed: {e}")
                if "429" in err_msg or "quota" in err_msg:
                    time.sleep(10)
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                else:
                    # Fall back to local rules for this chunk
                    print(f"[Gemini Categorize] Falling back to local rules for chunk {c_idx + 1}")
                    for item in chunk:
                        idx_val = item["index"]
                        categorized_items[idx_val] = {
                            "category": transactions[idx_val].get("category", "Other"),
                            "gst": transactions[idx_val].get("gst", ""),
                        }

    # Merge back into transactions
    for idx, txn in enumerate(transactions):
        cat_info = categorized_items.get(idx)
        if cat_info:
            txn["category"] = cat_info["category"]
            txn["gst"] = cat_info["gst"]

    return transactions, gemini_calls


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT FORMATTERS
# ══════════════════════════════════════════════════════════════════════════════

def clean_and_format_transactions(txns: List[Dict[str, Any]], date_format: str = "DD/MM/YYYY") -> List[Dict[str, Any]]:
    """
    pandas DataFrame cleanup: strip currency symbols, strip Dr/Cr suffixes,
    handle Indian lakh/crore format, fill NaN with empty string.
    """
    if not txns:
        return []

    df = pd.DataFrame(txns)
    df = df.fillna("")

    # Strip whitespace from string columns
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).str.strip()

    def clean_amount(val):
        if not val or val in ("—", "None", "null", "nil", "N/A", "n/a", "nan"):
            return ""
        cleaned = str(val).strip()
        # Handle parenthetical negatives: (500.00) → 500.00
        is_negative = False
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = cleaned[1:-1].strip()
            is_negative = True
        if cleaned.startswith('-'):
            cleaned = cleaned[1:].strip()
            is_negative = True
        # Remove currency symbols: ₹, Rs., INR, USD, $, £, €
        cleaned = re.sub(r'(?i)(rs\.?|inr|usd|₹|\$|£|€)', '', cleaned)
        # Remove Dr/Cr suffix (case-insensitive)
        cleaned = re.sub(r'(?i)\s*(dr|cr)\s*$', '', cleaned)
        # Remove any whitespace within the number (e.g., "10, 000.00" → "10,000.00")
        cleaned = re.sub(r'\s+', '', cleaned)
        # Keep only digits, decimal point, comma (thousands sep)
        cleaned = re.sub(r'[^\d\.,]', '', cleaned)
        # Remove commas (Indian lakh format: 1,00,000.00 → 100000.00)
        cleaned = cleaned.replace(',', '')
        # Collapse multiple dots — keep only integer.fraction
        if cleaned.count('.') > 1:
            parts = cleaned.split('.')
            cleaned = parts[0] + '.' + parts[-1]
        result = cleaned.strip()
        if result in ('', '.'):
            return ""
        # Validate it's actually a number
        try:
            float(result)
        except ValueError:
            return ""
        return result

    df['debit'] = df['debit'].apply(clean_amount)
    df['credit'] = df['credit'].apply(clean_amount)
    df['balance'] = df['balance'].apply(clean_amount)

    if 'description' in df.columns:
        # Clean up multi-line descriptions
        df['description'] = df['description'].apply(
            lambda x: re.sub(r'\s+', ' ', str(x)).strip()
        )

    # Ensure value_date column exists
    if 'value_date' not in df.columns:
        df['value_date'] = df['date'] if 'date' in df.columns else ""

    return df.to_dict(orient="records")


def generate_excel_file(txns: List[Dict[str, Any]], file_path: str):
    """
    Styled openpyxl Excel with purple (#4F46E5) header,
    zebra rows, auto column width, right-aligned amounts with #,##0.00 format.
    """
    if not txns:
        return

    df = pd.DataFrame(txns)

    cols = ["date", "value_date", "description", "debit", "credit", "balance", "category", "gst"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]

    rename_dict = {
        "date": "Date",
        "value_date": "Value Date",
        "description": "Description",
        "debit": "Debit",
        "credit": "Credit",
        "balance": "Balance",
        "category": "Category",
        "gst": "GST",
    }
    df = df.rename(columns=rename_dict)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Transactions", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Transactions"]

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        border_side = Side(border_style="thin", color="E2E8F0")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        cell_font = Font(name="Segoe UI", size=10)

        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
            cell.border = thin_border

        for row_idx in range(2, worksheet.max_row + 1):
            is_zebra = (row_idx % 2 == 0)
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border

                if is_zebra:
                    cell.fill = zebra_fill

                col_name = df.columns[col_idx - 1]
                if col_name in ["Debit", "Credit", "Balance"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    try:
                        val_str = str(cell.value).replace(",", "")
                        if val_str and val_str not in ("—", "None", ""):
                            cell.value = float(val_str)
                            cell.number_format = '#,##0.00'
                    except ValueError:
                        pass
                elif col_name == "Date":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto column width
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        worksheet.row_dimensions[1].height = 28
        for r in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[r].height = 20


def generate_csv_file(txns: List[Dict[str, Any]], file_path: str):
    """Generate CSV with utf-8-sig encoding for Excel compatibility."""
    if not txns:
        return
    df = pd.DataFrame(txns)
    cols = ["date", "value_date", "description", "debit", "credit", "balance", "category", "gst"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]
    df.to_csv(file_path, index=False, encoding='utf-8-sig')


# ══════════════════════════════════════════════════════════════════════════════
# MASTER FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

def parse_bank_statement_smart(
    pdf_path: str,
    password: str = None,
    categorize: bool = True,
    gst: bool = True,
    use_gemini_categorize: bool = False,
) -> Dict[str, Any]:
    """
    Master parsing function with the hybrid flow:

    1. check_pdf_basic() — read first 2 pages only, detect text vs scanned
    2. SCANNED → Gemini File API directly (1 API call, handles OCR)
    3. TEXT-BASED → native pdfplumber table extraction (0 API calls)
    4. Native fails → line-by-line regex parser (0 API calls)
    5. Native < 5 txns → is_text_quality_sufficient()
    6. Quality OK → Gemini text chunks (1-2 API calls)
    7. Quality poor → Gemini File API (1 API call)

    Args:
        pdf_path: Path to the PDF file
        password: PDF password if encrypted
        categorize: Whether to categorize transactions (True by default)
        gst: Whether to extract GST info (True by default)
        use_gemini_categorize: Use Gemini for categorization (False by default, uses local rules)

    Returns:
        {
            transactions: [...],
            method: "native" | "gemini_text" | "gemini_file",
            count: int,
            gemini_calls: int,
        }
    """
    overall_start = time.time()
    gemini_calls = 0
    method = "native"

    # ── Step 1: Basic PDF check ──
    print(f"\n{'=' * 60}")
    print(f"[Smart Parse] Starting: {pdf_path}")
    print(f"{'=' * 60}")

    total_pages, is_text_based = check_pdf_basic(pdf_path, password)
    print(f"[Smart Parse] Pages: {total_pages}, Text-based: {is_text_based}")

    if total_pages == 0:
        elapsed = time.time() - overall_start
        print(f"[Smart Parse] Empty PDF. Time: {elapsed:.2f}s")
        return {"transactions": [], "method": "native", "count": 0, "gemini_calls": 0}

    transactions = []

    # ── Step 2: SCANNED → Gemini File API directly ──
    if not is_text_based:
        print("[Smart Parse] Scanned PDF detected → Gemini File API (OCR)")
        method = "gemini_file"
        transactions, g_calls = parse_file_directly_with_gemini(
            pdf_path,
            "application/pdf",
            categorize=(categorize and use_gemini_categorize),
            gst=gst,
        )
        gemini_calls += g_calls

    else:
        # ── Step 3: TEXT-BASED → try native pdfplumber table extraction ──
        print("[Smart Parse] Text-based PDF → trying native pdfplumber extraction...")
        native_start = time.time()
        transactions = parse_pdf_natively(pdf_path, password)
        native_elapsed = time.time() - native_start
        print(f"[Smart Parse] Native extraction: {len(transactions)} txns in {native_elapsed:.2f}s")

        # ── Step 4: Native table parse fails → try line-by-line regex ──
        if not transactions:
            print("[Smart Parse] Native table extraction failed → trying line-by-line regex...")
            try:
                full_text = extract_full_text(pdf_path, password)
                transactions = parse_line_by_line(full_text)
                print(f"[Smart Parse] Line-by-line regex: {len(transactions)} txns")
            except Exception as e:
                print(f"[Smart Parse] Line-by-line regex failed: {e}")

        # ── Step 5: Native gets < 5 transactions → check text quality ──
        if len(transactions) < 5:
            print(f"[Smart Parse] Only {len(transactions)} native txns (< 5) → checking text quality...")

            full_text = extract_full_text(pdf_path, password)

            # ── Step 6: Text quality OK → Gemini text chunks ──
            if is_text_quality_sufficient(full_text):
                print("[Smart Parse] Text quality OK → Gemini text chunks")
                method = "gemini_text"
                transactions, g_calls = parse_with_gemini(
                    full_text,
                    categorize=(categorize and use_gemini_categorize),
                    gst=gst,
                )
                gemini_calls += g_calls
            else:
                # ── Step 7: Text quality poor → Gemini File API ──
                print("[Smart Parse] Text quality poor → Gemini File API")
                method = "gemini_file"
                transactions, g_calls = parse_file_directly_with_gemini(
                    pdf_path,
                    "application/pdf",
                    categorize=(categorize and use_gemini_categorize),
                    gst=gst,
                )
                gemini_calls += g_calls

    # ── Local categorization (FREE, default) ──
    if categorize and not use_gemini_categorize:
        # Apply local rule-based categorization
        for txn in transactions:
            desc = txn.get("description", "")
            txn["category"] = categorize_locally(desc)
            if gst:
                txn["gst"] = extract_gst_locally(desc)
            else:
                txn["gst"] = ""
    elif categorize and use_gemini_categorize and method == "native":
        # Native parse doesn't include Gemini categories — enrich with Gemini
        print("[Smart Parse] Enriching native results with Gemini categorization...")
        transactions, g_calls = categorize_and_tag_with_gemini(
            transactions,
            categorize=categorize,
            gst=gst,
        )
        gemini_calls += g_calls

    # ── Clean and format ──
    transactions = clean_and_format_transactions(transactions)

    # ── Deduplicate ──
    transactions = _deduplicate_transactions(transactions)

    elapsed = time.time() - overall_start
    print(f"\n{'=' * 60}")
    print(f"[Smart Parse] Done: {len(transactions)} txns via {method}")
    print(f"[Smart Parse] Gemini API calls: {gemini_calls}")
    print(f"[Smart Parse] Total time: {elapsed:.2f}s")
    print(f"{'=' * 60}\n")

    return {
        "transactions": transactions,
        "method": method,
        "count": len(transactions),
        "gemini_calls": gemini_calls,
    }
